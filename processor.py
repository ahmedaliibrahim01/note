import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def process_excels(shift_path, note_path=None, start_date=None, end_date=None):
    """
    Vardiya ve Note dosyalarını işler, çok sheet'li bir Excel çıktısı üretir.
    
    Sheet 1: Daily - Sorunlu kayıtlar (No schedule & No note)
    Sheet 2: Vardiya verileri (işlenmiş ve filtrelenmiş)
    Sheet 3: Note verileri (işlenmiş ve filtrelenmiş)
    Sheet 4: UserCodes - User, Level, Code ve Price bilgileri
    Sheet 5: Full Students - Vardiya ve Note karşılaştırması (filtrelenmiş)
    Sheet 6+: Her öğrenci için ayrı sheet
    
    Tarih filtreleme:
    - Eğer start_date ve end_date verilirse → Sadece o aralık
    - Eğer verilmezse → Tüm veriler
    """
    
    # Sheet 2: Vardiya İşleme
    df_shift = process_shift_excel(shift_path)
    
    # Sheet 3: Note İşleme (eğer varsa)
    df_note = None
    if note_path:
        df_note = process_note_excel(note_path)
    
    # TARİH FİLTRELEME (eğer tarih aralığı seçilmişse)
    if start_date and end_date:
        # Tarihleri datetime'a çevir
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
        
        # Sıralama kontrolü (kullanıcı ters seçmiş olabilir)
        if start > end:
            start, end = end, start
            print("⚠️ Tarihler ters sırada, otomatik düzeltildi.")
        
        print(f"\n📅 Filtreleme: {start.strftime('%m/%d/%Y')} - {end.strftime('%m/%d/%Y')}")
        
        # Vardiya'yı filtrele
        df_shift_filtered = df_shift[
            (df_shift['Service Date'] >= start) & 
            (df_shift['Service Date'] <= end)
        ].copy()
        
        if len(df_shift_filtered) > 0:
            df_shift = df_shift_filtered.reset_index(drop=True)
            print(f"✅ Vardiya: {len(df_shift)} kayıt filtrelendi")
        else:
            print("⚠️ Vardiya: Bu tarih aralığında veri yok, tüm veriler gösteriliyor")
        
        # Note'u filtrele (eğer varsa)
        if df_note is not None:
            df_note_filtered = df_note[
                (df_note['Service Date'] >= start) & 
                (df_note['Service Date'] <= end)
            ].copy()
            
            if len(df_note_filtered) > 0:
                df_note = df_note_filtered.reset_index(drop=True)
                print(f"✅ Note: {len(df_note)} kayıt filtrelendi")
            else:
                print("⚠️ Note: Bu tarih aralığında veri yok, tüm veriler gösteriliyor")
    else:
        print("\n📊 Filtreleme yok, tüm veriler işleniyor")
    
    # Sheet 4: UserCodes
    df_usercodes = process_usercodes()
    
    # Sheet 5: Full Students (eğer Note varsa, filtrelenmiş verilerle)
    df_full_students = None
    df_daily = None
    if df_note is not None and len(df_note) > 0:
        df_full_students = create_full_students(df_shift, df_note)
        print(f"✅ Full Students: {len(df_full_students)} kayıt oluşturuldu")
        
        # Sheet 1: Daily - Sorunlu kayıtlar (Full Students'tan)
        df_daily = create_daily_sheet(df_full_students)
        if len(df_daily) > 0:
            print(f"✅ Daily: {len(df_daily)} sorunlu kayıt bulundu")
        else:
            print("✅ Daily: Sorunlu kayıt yok (tüm kayıtlar eşleşmiş)")
    else:
        print("⚠️ Note verisi yok, Full Students ve Daily oluşturulamadı")
    
    # Excel dosyasını oluştur
    output_path = "uploads/final_report.xlsx"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Daily (sorunlu kayıtlar, eğer varsa)
        if df_daily is not None:
            df_daily.to_excel(writer, sheet_name='Daily', index=False)
        
        # Sheet 2: Vardiya (filtrelenmiş veya tüm veriler)
        df_shift.to_excel(writer, sheet_name='Vardiya', index=False)
        
        # Sheet 3: Note (filtrelenmiş veya tüm veriler, eğer varsa)
        if df_note is not None and len(df_note) > 0:
            df_note.to_excel(writer, sheet_name='Note', index=False)
        
        # Sheet 4: UserCodes
        df_usercodes.to_excel(writer, sheet_name='UserCodes', index=False)
        
        # Sheet 5: Full Students (filtrelenmiş veya tüm veriler, eğer varsa)
        if df_full_students is not None and len(df_full_students) > 0:
            df_full_students.to_excel(writer, sheet_name='Full Students', index=False)
    
    # Formüller ve formatları ekle
    if df_note is not None and len(df_note) > 0:
        add_formulas_and_formatting(output_path)
    
    # Her öğrenci için ayrı sheet oluştur
    if df_full_students is not None and len(df_full_students) > 0:
        create_student_sheets(output_path)
    
    print(f"\n✅ Excel dosyası oluşturuldu: {output_path}\n")
    
    return output_path


def create_full_students(df_vardiya, df_note):
    """
    Vardiya ve Note sheet'lerini karşılaştırarak Full Students sheet'ini oluşturur.
    
    Mantık:
    1. Note'taki tüm satırları ekle
       - Vardiya'da eşleşen var → Status: "", Code: Vardiya'dan
       - Vardiya'da eşleşen yok → Status: "No schedule", Code: Boş
    2. Vardiya'daki eşleşmeyenleri ekle
       - Status: "No note", Code: Vardiya'dan
    
    Eşleştirme: Student + Service Date + User (case-insensitive)
    """
    
    # Tarihleri normalize et (ikisi de datetime olsun)
    df_vardiya = df_vardiya.copy()
    df_note = df_note.copy()
    
    df_vardiya['Service Date'] = pd.to_datetime(df_vardiya['Service Date'], errors='coerce').dt.normalize()
    df_note['Service Date'] = pd.to_datetime(df_note['Service Date'], errors='coerce').dt.normalize()
    
    # Vardiya anahtarlarını oluştur (eşleştirme için)
    vardiya_keys = {}
    for idx, row in df_vardiya.iterrows():
        student = str(row['Student']).strip().lower()
        date = row['Service Date']
        user = str(row['User']).strip().lower()
        
        key = (student, date, user)
        vardiya_keys[key] = idx
    
    # Full Students verilerini topla
    full_students_data = []
    matched_vardiya_keys = set()
    
    # ADIM 1: Note'taki tüm satırları işle
    for idx, row in df_note.iterrows():
        student = str(row['Student']).strip().lower()
        date = row['Service Date']
        user = str(row['User']).strip().lower()
        
        key = (student, date, user)
        
        # Vardiya'da bu key var mı?
        if key in vardiya_keys:
            status = ""
            matched_vardiya_keys.add(key)
            
            # Code'u Vardiya'dan al
            vardiya_idx = vardiya_keys[key]
            vardiya_row = df_vardiya.iloc[vardiya_idx]
            code = vardiya_row['Code']
        else:
            status = "No schedule"
            code = ""
        
        # Note satırını ekle
        row_data = [
            row['Student'],
            row['Service Date'],
            row['Session Time'],
            row['Session Duration'],
            row['Hours'],
            row['Pay'],
            row['Units'],
            row['FNF'],
            code,
            row['Charged Amount'],
            row['User'],
            row['Level'],
            status,
            row['Unbilled notes'],
            row['Expiration Date']
        ]
        
        full_students_data.append(row_data)
    
    # ADIM 2: Vardiya'daki eşleşmeyenleri işle
    for idx, row in df_vardiya.iterrows():
        student = str(row['Student']).strip().lower()
        date = row['Service Date']
        user = str(row['User']).strip().lower()
        
        key = (student, date, user)
        
        if key not in matched_vardiya_keys:
            status = "No note"
            
            row_data = [
                row['Student'],
                row['Service Date'],
                row['Session Time'],
                row['Session Duration'],
                row['Hours'],
                row['Pay'],
                row['Units'],
                row['FNF'],
                row['Code'],
                row['Charged Amount'],
                row['User'],
                row['Level'],
                status,
                row['Unbilled notes'],
                row['Expiration Date']
            ]
            
            full_students_data.append(row_data)
    
    # DataFrame oluştur
    df_full_students = pd.DataFrame(full_students_data, columns=[
        'Student',
        'Service Date',
        'Session Time',
        'Session Duration',
        'Hours',
        'Pay',
        'Units',
        'FNF',
        'Code',
        'Charged Amount',
        'User',
        'Level',
        'Note',
        'Unbilled notes',
        'Expiration Date'
    ])
    
    # Service Date'i datetime'a çevir
    df_full_students['Service Date'] = pd.to_datetime(df_full_students['Service Date'], errors='coerce')
    
    # Tarihe göre sırala
    df_full_students = df_full_students.sort_values('Service Date', na_position='last')
    
    return df_full_students


def create_daily_sheet(df_full_students):
    """
    Full Students'tan sorunlu kayıtları (No schedule & No note) filtreler.
    
    Özellikler:
    - Sadece "No schedule" ve "No note" kayıtları
    - Öğrenciye göre alfabetik sıralı
    - Tüm sütunlar Full Students ile aynı
    """
    
    # Sorunlu kayıtları filtrele (Note sütunu boş OLMAYANLAR)
    df_daily = df_full_students[
        (df_full_students['Note'] == 'No schedule') | 
        (df_full_students['Note'] == 'No note')
    ].copy()
    
    # Öğrenciye göre alfabetik sırala
    df_daily = df_daily.sort_values('Student', na_position='last')
    
    # Index'i sıfırla
    df_daily = df_daily.reset_index(drop=True)
    
    return df_daily


def process_usercodes():
    """
    UserCodes sheet'i için User, Level, Code ve Price bilgilerini oluşturur.
    """
    users = [
        "Abduallahi, Ikraan",
        "Abdulle, Suweda",
        "Abdulle, Zamzam",
        "Abdurahman, Yasmin",
        "Ahmed, Jama",
        "ALI, AMIRA",
        "Ali, Naima",
        "Ali, Nasteho",
        "Barack, Sara",
        "Barqab, Mushtaq",
        "Bello, Hiqmat",
        "Hashi, Ridwan",
        "Jama, Sara",
        "Khalif, Muna",
        "Mohamed, Meymun",
        "Muse, Aaliya",
        "Muse, Adna",
        "Muse, Ayni",
        "Noor, Kamilo",
        "Omar, Saido",
        "Ran, Sameya",
        "Kounlavouth, Roselie",
        "El-Tay, Dr. Ismael",
        "Nate",
        "Osman, Basra",
        "Khalif, MK",
        "Hashi, Kafiya",
        "MUHUMED, YACQUB",
        "Omar, Mohamed",
        "Mohamed, Madar",
        "Issa, Yassin",
        "Sharif, Omar",
        "Jama, Ruweyda",
        "Hawadle, Abdiaziz",
        "Mohamed, Samsam",
        "Muse, Nawaal",
        "Mohamed, Mohamed",
        "Ibrahim, Ahmed",
        "Webi3, Amina"
    ]
    
    levels = [
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 1",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 1",
        "Level 2",
        "Level 2",
        "Level 1",
        "Level 1",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 1",
        "Level 1",
        "Level 2",
        "Level 1",
        "Level 1",
        "Level 1",
        "Level 1",
        "Level 1",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 1",
        "Level 2",
        "Level 2",
        "Level 1",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 2",
        "Level 2"
    ]
    
    codes = [
        "97151", 
        "PM", 
        "T1024", 
        "1/1 \n97153",
        "97154", 
        "0373T", 
        "O&D \n97155",
        "FT", 
        "97157", 
        "H0046"
    ]
    prices = [50.11, 94.80, 112.67, 20.17, 6.72, 24.19, 20.17, 20.17, 6.72, 0.52]
    
    max_rows = max(len(users), len(codes))
    
    df = pd.DataFrame({
        'User': users + [''] * (max_rows - len(users)),
        'Level': levels + [''] * (max_rows - len(levels)),
        'C': [''] * max_rows,
        'D': [''] * max_rows,
        'Code': codes + [''] * (max_rows - len(codes)),
        'Price': prices + [''] * (max_rows - len(prices))
    })
    
    return df


def create_student_sheets(file_path):
    """
    Full Students sheet'inden her öğrenci için ayrı sheet oluşturur.
    
    Özellikler:
    - Her öğrenci için ayrı sheet
    - Sheet ismi = Öğrenci ismi
    - Tüm sütunlar ve veriler kopyalanır
    - Formüller korunur
    - Tarih formatı korunur
    - Boş öğrenciler atlanır
    """
    wb = load_workbook(file_path)
    
    # Full Students sheet'i kontrol et
    if 'Full Students' not in wb.sheetnames:
        print("⚠️ Full Students sheet'i bulunamadı, öğrenci sheet'leri oluşturulamadı")
        wb.close()
        return
    
    ws_full = wb['Full Students']
    
    # Başlık satırını al (row 1)
    headers = []
    for cell in ws_full[1]:
        headers.append(cell.value)
    
    # Student sütununun index'ini bul
    try:
        student_col_idx = headers.index('Student')
    except ValueError:
        print("⚠️ Student sütunu bulunamadı")
        wb.close()
        return
    
    # Tüm öğrencileri topla ve grupla
    students_data = {}
    
    for row_idx in range(2, ws_full.max_row + 1):
        student_name = ws_full.cell(row=row_idx, column=student_col_idx + 1).value
        
        # Boş öğrenci isimlerini atla
        if not student_name or str(student_name).strip() == "":
            continue
        
        student_name = str(student_name).strip()
        
        # Öğrenci için ilk kez mi?
        if student_name not in students_data:
            students_data[student_name] = []
        
        # Bu satırı öğrenciye ekle
        students_data[student_name].append(row_idx)
    
    print(f"\n📊 {len(students_data)} farklı öğrenci bulundu")
    
    # Her öğrenci için sheet oluştur
    for student_name, row_indices in students_data.items():
        # Sheet ismini temizle (Excel yasak karakterleri)
        sheet_name = student_name
        for char in [':', '/', '\\', '?', '*', '[', ']']:
            sheet_name = sheet_name.replace(char, '-')
        
        # Sheet ismi 31 karakterden uzunsa kısalt
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        
        # Sheet oluştur
        ws_student = wb.create_sheet(title=sheet_name)
        
        # Başlık satırını kopyala
        for col_idx, header in enumerate(headers, start=1):
            ws_student.cell(row=1, column=col_idx, value=header)
        
        # Bu öğrencinin satırlarını kopyala
        new_row = 2
        for source_row_idx in row_indices:
            for col_idx in range(1, len(headers) + 1):
                source_cell = ws_full.cell(row=source_row_idx, column=col_idx)
                target_cell = ws_student.cell(row=new_row, column=col_idx)
                
                # Değeri kopyala
                target_cell.value = source_cell.value
                
                # Formül mü değer mi kontrol et
                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    # Formülü kopyala (satır numarasını güncelle)
                    formula = source_cell.value
                    # Formüldeki satır numarasını değiştir
                    old_row_ref = str(source_row_idx)
                    new_row_ref = str(new_row)
                    formula = formula.replace(f'D{old_row_ref}', f'D{new_row_ref}')
                    formula = formula.replace(f'E{old_row_ref}', f'E{new_row_ref}')
                    formula = formula.replace(f'H{old_row_ref}', f'H{new_row_ref}')
                    formula = formula.replace(f'I{old_row_ref}', f'I{new_row_ref}')
                    formula = formula.replace(f'K{old_row_ref}', f'K{new_row_ref}')
                    formula = formula.replace(f'L{old_row_ref}', f'L{new_row_ref}')
                    target_cell.value = formula
                
                # Tarih formatını kopyala (B sütunu - Service Date)
                if col_idx == 2:
                    target_cell.number_format = 'MM/DD/YYYY'
            
            new_row += 1
        
        print(f"✅ '{sheet_name}' sheet'i oluşturuldu ({len(row_indices)} kayıt)")
    
    # Kaydet ve kapat
    wb.save(file_path)
    wb.close()
    
    print(f"\n🎉 Tüm öğrenci sheet'leri oluşturuldu!\n")


def add_formulas_and_formatting(file_path):
    """
    Daily, Note, Vardiya ve Full Students sheet'lerine formüller ve formatlar ekler.
    
    Daily:
    - B sütunu tarih formatı
    - E, F, G, H, J, L sütunlarına formüller
    
    Note:
    - B sütunu tarih formatı
    - E, F, G, H, J, L sütunlarına formüller
    
    Vardiya:
    - B sütunu tarih formatı
    
    Full Students:
    - B sütunu tarih formatı
    - E, F, G, H, J, L sütunlarına formüller
    """
    wb = load_workbook(file_path)
    
    # Daily sheet'i işle
    if 'Daily' in wb.sheetnames:
        ws_daily = wb['Daily']
        row_count_daily = ws_daily.max_row
        
        # B sütununun tarih formatını ayarla
        for row in range(2, row_count_daily + 1):
            b_cell = ws_daily[f'B{row}']
            if b_cell.value:
                b_cell.number_format = 'MM/DD/YYYY'
        
        # Formülleri ekle
        for row in range(2, row_count_daily + 1):
            # E sütunu formülü - Hours
            e_cell = ws_daily[f'E{row}']
            e_cell.value = f'=HOUR(D{row})+MINUTE(D{row})/100'
            
            # F sütunu formülü - Pay
            f_cell = ws_daily[f'F{row}']
            f_cell.value = f'=ROUND(INT(E{row}) + (E{row} - INT(E{row})) * 100 / 60, 2)'
            
            # G sütunu formülü - Units
            g_cell = ws_daily[f'G{row}']
            g_cell.value = f'=E{row}*4'
            
            # H sütunu formülü - FNF
            h_cell = ws_daily[f'H{row}']
            h_cell.value = (
                f'=LET('
                f'total, INT(E{row})*60 + (E{row}-INT(E{row}))*100, '
                f'blocks, QUOTIENT(total, 15), '
                f'extra, MOD(total, 15), '
                f'units, blocks + IF(extra>=8, 1, 0), '
                f'units)'
            )
            
            # J sütunu formülü - Charged Amount
            j_cell = ws_daily[f'J{row}']
            j_cell.value = (
                f'=IF(L{row}="Level 1", '
                f'VLOOKUP(I{row}, UserCodes!E:F, 2, FALSE) * H{row}, '
                f'IF(L{row}="Level 2", '
                f'VLOOKUP(I{row}, UserCodes!E:F, 2, FALSE) * H{row} * 0.8, 0))'
            )
            
            # L sütunu formülü - Level
            l_cell = ws_daily[f'L{row}']
            l_cell.value = f'=VLOOKUP(K{row}, UserCodes!A:B, 2, FALSE)'
    
    # Note sheet'i işle
    if 'Note' in wb.sheetnames:
        ws = wb['Note']
        row_count = ws.max_row
        
        # B sütununun tarih formatını ayarla
        for row in range(2, row_count + 1):
            b_cell = ws[f'B{row}']
            if b_cell.value:
                b_cell.number_format = 'MM/DD/YYYY'
        
        # Formülleri ekle
        for row in range(2, row_count + 1):
            # E sütunu formülü
            e_cell = ws[f'E{row}']
            e_cell.value = f'=HOUR(D{row})+MINUTE(D{row})/100'
            
            # F sütunu formülü
            f_cell = ws[f'F{row}']
            f_cell.value = f'=ROUND(INT(E{row}) + (E{row} - INT(E{row})) * 100 / 60, 2)'
            
            # G sütunu formülü
            g_cell = ws[f'G{row}']
            g_cell.value = f'=E{row}*4'
            
            # H sütunu formülü
            h_cell = ws[f'H{row}']
            h_cell.value = (
                f'=LET('
                f'total, INT(E{row})*60 + (E{row}-INT(E{row}))*100, '
                f'blocks, QUOTIENT(total, 15), '
                f'extra, MOD(total, 15), '
                f'units, blocks + IF(extra>=8, 1, 0), '
                f'units)'
            )
            
            # J sütunu formülü
            j_cell = ws[f'J{row}']
            j_cell.value = (
                f'=IF(L{row}="Level 1", '
                f'VLOOKUP(I{row}, UserCodes!E:F, 2, FALSE) * H{row}, '
                f'IF(L{row}="Level 2", '
                f'VLOOKUP(I{row}, UserCodes!E:F, 2, FALSE) * H{row} * 0.8, 0))'
            )
            
            # L sütunu formülü
            l_cell = ws[f'L{row}']
            l_cell.value = f'=VLOOKUP(K{row}, UserCodes!A:B, 2, FALSE)'
    
    # Vardiya sheet'i işle
    if 'Vardiya' in wb.sheetnames:
        ws_vardiya = wb['Vardiya']
        row_count_vardiya = ws_vardiya.max_row
        
        # B sütununun tarih formatını ayarla
        for row in range(2, row_count_vardiya + 1):
            b_cell = ws_vardiya[f'B{row}']
            if b_cell.value:
                b_cell.number_format = 'MM/DD/YYYY'
    
    # Full Students sheet'i işle
    if 'Full Students' in wb.sheetnames:
        ws_full = wb['Full Students']
        row_count_full = ws_full.max_row
        
        # B sütununun tarih formatını ayarla
        for row in range(2, row_count_full + 1):
            b_cell = ws_full[f'B{row}']
            if b_cell.value:
                b_cell.number_format = 'MM/DD/YYYY'
        
        # Formülleri ekle
        for row in range(2, row_count_full + 1):
            # E sütunu formülü
            e_cell = ws_full[f'E{row}']
            e_cell.value = f'=HOUR(D{row})+MINUTE(D{row})/100'
            
            # F sütunu formülü
            f_cell = ws_full[f'F{row}']
            f_cell.value = f'=ROUND(INT(E{row}) + (E{row} - INT(E{row})) * 100 / 60, 2)'
            
            # G sütunu formülü
            g_cell = ws_full[f'G{row}']
            g_cell.value = f'=E{row}*4'
            
            # H sütunu formülü
            h_cell = ws_full[f'H{row}']
            h_cell.value = (
                f'=LET('
                f'total, INT(E{row})*60 + (E{row}-INT(E{row}))*100, '
                f'blocks, QUOTIENT(total, 15), '
                f'extra, MOD(total, 15), '
                f'units, blocks + IF(extra>=8, 1, 0), '
                f'units)'
            )
            
            # J sütunu formülü
            j_cell = ws_full[f'J{row}']
            j_cell.value = (
                f'=IF(L{row}="Level 1", '
                f'VLOOKUP(I{row}, UserCodes!E:F, 2, FALSE) * H{row}, '
                f'IF(L{row}="Level 2", '
                f'VLOOKUP(I{row}, UserCodes!E:F, 2, FALSE) * H{row} * 0.8, 0))'
            )
            
            # L sütunu formülü
            l_cell = ws_full[f'L{row}']
            l_cell.value = f'=VLOOKUP(K{row}, UserCodes!A:B, 2, FALSE)'
    
    wb.save(file_path)
    wb.close()


def process_shift_excel(shift_path):
    """Vardiya dosyasını işler"""
    df = pd.read_excel(shift_path)

    # D, E, I-W sütunlarını sil
    cols_to_remove = [3,4] + list(range(8,23))
    cols_to_remove_existing = [i for i in cols_to_remove if i < len(df.columns)]
    df = df.drop(df.columns[cols_to_remove_existing], axis=1)

    # E sütunu değer güncelleme
    if len(df.columns) >= 5:
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdulkadir, Nadra" if str(x) == "NaAb" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Mohamed, Sumaya" if str(x) == "SuMo" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Yusuf, Khadija" if str(x) == "Kh.I.Y" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Yusuf 2, Khadra" if str(x) == "Kh.R.Y" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Kadir, Hamdi" if str(x) == "HaKa" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdi, Ahmed" if str(x) == "AhAB" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdi, Mohamed" if str(x) == "MoAB" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdi, Safia" if str(x) == "SaAb" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdi, Salman" if str(x) == "Sal.Ab" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Ali, Saami" if str(x) == "Saa.Al" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Sulub, Sharmake" if str(x) == "SHSU" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdi, Samira" if str(x) == "SamAb" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Jama, Khadija" if str(x) == "K.Jam" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Abdullahi, Abdirahman" if str(x) == "AbAb" else x)
        df.iloc[:,4] = df.iloc[:,4].apply(lambda x: "Omar, Abdiweli" if str(x) == "AbOm" else x)

    # F sütunu "Last, First" formatına çevirme
    if len(df.columns) >= 6:
        def reverse_name(x):
            parts = str(x).split()
            if len(parts) == 2:
                return f"{parts[1]}, {parts[0]}"
            else:
                return x
        df.iloc[:,5] = df.iloc[:,5].apply(reverse_name)

    # B ve C sütunlarını birleştir
    if len(df.columns) >= 3:
        df.iloc[:,1] = df.iloc[:,1].astype(str) + " - " + df.iloc[:,2].astype(str)
        df.iloc[:,2] = ""

    # A-F sütunlarını P-U sütunlarına kopyala
    for i in range(9):
        df[f'Empty_{i}'] = ""
    
    for i in range(6):
        col_name = f'Copy_{df.columns[i]}'
        df[col_name] = df.iloc[:, i]
    
    for i in range(6):
        df.iloc[:, i] = ""
        df.columns.values[i] = "Empty"
    
    # Taşıma işlemleri
    if len(df.columns) >= 20:
        df.iloc[:, 0] = df.iloc[:, 19]
        df.columns.values[0] = "Student"
    
    if len(df.columns) >= 16:
        df.iloc[:, 1] = df.iloc[:, 15]
        df.columns.values[1] = "Service Date"
    
    if len(df.columns) >= 17:
        df.iloc[:, 2] = df.iloc[:, 16]
        df.columns.values[2] = "Session Time"
    
    if len(df.columns) >= 19:
        df.iloc[:, 8] = df.iloc[:, 18]
        
        # I sütunu (Code) değerlerini değiştir
        def format_code(x):
            x_str = str(x)
            if "97153 | Intervention 1-on-1" in x_str or "97153|Intervention 1-on-1" in x_str:
                return "1/1 \n97153"
            elif "97155 | O&D Supervision" in x_str or "97155|O&D Supervision" in x_str:
                return "O&D \n97155"
            else:
                return x
        
        df.iloc[:, 8] = df.iloc[:, 8].apply(format_code)
    
    if len(df.columns) >= 21:
        df.iloc[:, 10] = df.iloc[:, 20]
    
    # Başlıkları ayarla
    if len(df.columns) >= 15:
        df.columns.values[3] = "Session Duration"
        df.columns.values[4] = "Hours"
        df.columns.values[5] = "Pay"
        df.columns.values[6] = "Units"
        df.columns.values[7] = "FNF"
        df.columns.values[8] = "Code"
        df.columns.values[9] = "Charged Amount"
        df.columns.values[10] = "User"
        df.columns.values[11] = "Level"
        df.columns.values[12] = "Note"
        df.columns.values[13] = "Unbilled notes"
        df.columns.values[14] = "Expiration Date"
    
    # P-U sütunlarını sil
    if len(df.columns) >= 21:
        df = df.drop(df.columns[15:21], axis=1)

    # Service Date'i datetime'a çevir
    if 'Service Date' in df.columns:
        df['Service Date'] = pd.to_datetime(df['Service Date'], errors='coerce').dt.normalize()

    return df


def process_note_excel(note_path):
    """
    Note dosyasını işler.
    """
    df = pd.read_excel(note_path)
    
    # B sütunu tarih formatını düzelt
    if len(df.columns) >= 2:
        df.iloc[:, 1] = pd.to_datetime(df.iloc[:, 1], errors='coerce').dt.normalize()
    
    # F-L sütunlarını sil
    cols_to_remove = list(range(5, 12))
    cols_to_remove_existing = [i for i in cols_to_remove if i < len(df.columns)]
    
    if cols_to_remove_existing:
        df = df.drop(df.columns[cols_to_remove_existing], axis=1)
    
    # E-J arası boş sütunlar oluştur
    for i in range(6):
        df[f'Empty_{i}'] = ""
    
    # E sütunundaki User verilerini geçici olarak kaydet
    df['User_Temp'] = df.iloc[:, 4]
    
    # E sütununu boşalt
    df.iloc[:, 4] = ""
    
    # User_Temp'i K sütununa taşı
    if len(df.columns) >= 12:
        df.iloc[:, 10] = df.iloc[:, 11]
        df = df.drop(df.columns[11], axis=1)
    
    # L-O arası 4 boş sütun ekle
    for i in range(4):
        df[f'Empty_Extra_{i}'] = ""
    
    # A-O başlıklarını ayarla
    if len(df.columns) >= 15:
        df.columns.values[0] = "Student"
        df.columns.values[1] = "Service Date"
        df.columns.values[2] = "Session Time"
        df.columns.values[3] = "Session Duration"
        df.columns.values[4] = "Hours"
        df.columns.values[5] = "Pay"
        df.columns.values[6] = "Units"
        df.columns.values[7] = "FNF"
        df.columns.values[8] = "Code"
        df.columns.values[9] = "Charged Amount"
        df.columns.values[10] = "User"
        df.columns.values[11] = "Level"
        df.columns.values[12] = "Note"
        df.columns.values[13] = "Unbilled notes"
        df.columns.values[14] = "Expiration Date"
    
    return df